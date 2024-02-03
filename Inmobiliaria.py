from notifypy import Notify
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta


def leer_o_crear_excel(nombre_archivo):
    notifiaction = Notify()


    notifiaction.message="alerta de vencimiento"
    try:
        # Intenta abrir el archivo Excel existente
        wb = openpyxl.load_workbook(nombre_archivo)

    except FileNotFoundError:
        # Si el archivo no existe, crea uno nuevo
        notifiaction.application_name="Recordatorio app"

        notifiaction.title=f"El archivo {nombre_archivo} no existe. Se ha creado un nuevo archivo."
        notifiaction.send()

        wb = Workbook()
        wb.save(nombre_archivo)
        print(f"El archivo {nombre_archivo} no existe. Se ha creado un nuevo archivo.")

        ws = wb.active
        encabezados = ["Nombre", 
                       "Fecha inicial",
                       "Meses para vencimiento",
                       "Fecha Vencimiento",
                       "Meses para actualizacion",
                         "Fecha actualizacion",
                         "Meses hasta el vencimiento",
                         "Meses hasta la actualizacion"]
        for col, encabezado in enumerate(encabezados, start=1):
            # Obtiene la letra de la columna
            letra_columna = get_column_letter(col)
            # Escribe el encabezado en la celda correspondiente
            ws[f"{letra_columna}1"] = encabezado
            wb.save(nombre_archivo)
    else:
        notifiaction = Notify()
        notifiaction.application_name="Recordatorio app"

        notifiaction.title="Leyendo excel..."
        notifiaction.message="se esta leyendo el excel"
        notifiaction.icon = "bottaro.png"
        notifiaction.send()
        ws = wb.active
        ws = CalcularFechasVencimiento(ws)
        wb.save(nombre_archivo) 
         

def CalcularFechasVencimiento(ws):
    notifiaction = Notify()

    indice_fecha = None
    indice_periodo_vencimiento = None
    indice_Fecha_vencimiento = None
    indice_LeftMonths_vencimiento = None
    indice_LeftMonths_renovacion = None
    indice_Fecha_renovacion = None
    indice_periodo_renovacion = None
    indice_nombre = None
    for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == "Nombre":
                    indice_nombre = col
                elif ws.cell(row=1, column=col).value == "Fecha inicial":
                    indice_fecha = col
                elif ws.cell(row=1, column=col).value == "Meses para vencimiento":
                    indice_periodo_vencimiento = col
                elif ws.cell(row=1, column=col).value == "Meses para actualizacion":
                    indice_periodo_renovacion = col
                elif ws.cell(row=1, column=col).value == "Fecha Vencimiento":
                    indice_Fecha_vencimiento = col
                elif ws.cell(row=1, column=col).value == "Fecha actualizacion":
                    indice_Fecha_renovacion = col
                elif ws.cell(row=1, column=col).value == "Meses hasta el vencimiento":
                    indice_LeftMonths_vencimiento = col
                elif ws.cell(row=1, column=col).value == "Meses hasta la actualizacion":
                    indice_LeftMonths_renovacion = col


    if indice_fecha is None or indice_periodo_vencimiento is None:
            notifiaction.title = "No se encontraron las columnas 'Fecha inicial' y 'Periodo de Vencimiento'."
            notifiaction.send()
            print("No se encontraron las columnas 'Fecha inicial' y 'Periodo de Vencimiento'.")
            return
    # Calcula la fecha de vencimiento y escribe en la columna "Fecha de Vencimiento"
    for row in range(2, ws.max_row + 1):
        fecha = ws.cell(row=row, column=indice_fecha).value
        periodo_vencimiento = ws.cell(row=row, column=indice_periodo_vencimiento).value
        periodo_renovacion = ws.cell(row=row, column=indice_periodo_renovacion).value

        if fecha:
            if fecha is str:
                fecha = datetime.strptime(fecha, "%d/%m/%Y")
            if periodo_vencimiento:   
                if (periodo_vencimiento>0):
                    fecha_vencimiento = fecha + relativedelta(months=periodo_vencimiento)
            if periodo_renovacion:
                 if (periodo_renovacion>0):
                    fecha_renovacion = fecha + relativedelta(months=periodo_renovacion)
                    
                
            fecha_hoy = datetime.now()

            rd = relativedelta(years=0, months=0, days=0)

            fecha_modificada =  rd + fecha_hoy

            left_months_vencimiento = relativedelta(fecha_vencimiento, fecha_modificada)
            meses_restantes_vencimiento = left_months_vencimiento.months+(left_months_vencimiento.years*12)+ left_months_vencimiento.weeks//3

            left_months_renovacion = relativedelta(fecha_renovacion, fecha_modificada)
            meses_restantes_renovacion = left_months_renovacion.months+(left_months_renovacion.years*12) + left_months_renovacion.weeks//3

            if(meses_restantes_vencimiento<=1 and meses_restantes_renovacion<=1):
                notifiaction = Notify()
                notifiaction.icon = "bottaro.png"
                notifiaction.application_name="Recordatorio app"
                notifiaction.title = "Alerta: Vencimiento y renovacion inminente"
                notifiaction.message = f"{ws.cell(row=row, column=indice_nombre).value} en la fila {row} le quedan {meses_restantes_vencimiento} mes restantes"
                notifiaction.send()
            elif(meses_restantes_vencimiento<=1):
                notifiaction = Notify()
                notifiaction.application_name="Recordatorio app"

                notifiaction.icon = "bottaro.png"

                notifiaction.title = "Alerta: Vencimiento inminente"
                notifiaction.message = f"{ws.cell(row=row, column=indice_nombre).value} en la fila {row} le quedan {meses_restantes_vencimiento} mes restantes"
                notifiaction.send()

            elif(meses_restantes_renovacion<=1):
                notifiaction = Notify()
                notifiaction.application_name="Recordatorio app"
                notifiaction.icon = "bottaro.png"

                notifiaction.title="Alerta: Renovacion inminente"
                notifiaction.message= f"{ws.cell(row=row, column=indice_nombre).value} en la fila {row} le quedan {meses_restantes_renovacion} mes restantes"
                notifiaction.send()

            #indice_nombre
            ws.cell(row=row, column=indice_LeftMonths_renovacion).value = meses_restantes_renovacion
            ws.cell(row=row, column=indice_Fecha_renovacion).value = fecha_renovacion.strftime("%d/%m/%Y")

            ws.cell(row=row, column=indice_LeftMonths_vencimiento).value = meses_restantes_vencimiento
            ws.cell(row=row, column=indice_Fecha_vencimiento).value = fecha_vencimiento.strftime("%d/%m/%Y")
    return ws

leer_o_crear_excel("Vencimientos.xlsx")
